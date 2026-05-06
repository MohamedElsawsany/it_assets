"""
it_assets/export_utils.py
─────────────────────────
Shared helpers for exporting querysets to Excel (.xlsx) or PDF.
"""

import io


def export_xlsx(filename, headers, rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename.capitalize()
    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(color='FFFFFF', bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}.xlsx"'},
    )


def export_pdf(title, headers, rows, landscape=False):
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from django.http import HttpResponse

    buf = io.BytesIO()
    pagesize = rl_landscape(A4) if landscape else A4
    doc = SimpleDocTemplate(buf, pagesize=pagesize, leftMargin=30, rightMargin=30,
                            topMargin=40, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(title, styles['Title']),
        Spacer(1, 12),
    ]
    table_data = [headers] + rows
    col_count = len(headers)
    page_width = pagesize[0] - 60
    col_width = page_width / col_count
    t = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',   (0, 0), (-1, 0), colors.white),
        ('FONTNAME',    (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0), 9),
        ('FONTSIZE',    (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F8')]),
        ('GRID',        (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',  (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return HttpResponse(
        buf.read(),
        content_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{title.lower()}.pdf"'},
    )
